#!/usr/bin/env python3
"""
build_cvh_data_download.py
Generate TBL-Data-CVH-202603.xlsx for the CVH Working Americans' Tax Cut Act analysis.
Figures 1–7 and Table 1, matching the Booker template format.
"""

import csv
import os
import sys
sys.path.insert(0, '/home/jar335/.local/lib/python3.12/site-packages')

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
VINTAGE  = "202603120647"
ROOT     = f"/nfs/roberts/scratch/pi_nrs36/jar335/model_data/Tax-Simulator/v1/{VINTAGE}"
GDP_PATH = "/nfs/roberts/project/pi_nrs36/shared/model_data/Macro-Projections/v3/2026022522/baseline/projections.csv"
OUT      = os.path.join(ROOT, "TBL-Data-CVH-202603.xlsx")

YEAR_SHOW = 2026
SCENARIOS = ["alt_max", "surtax"]
PROVISIONS = ["Alternative Maximum Tax", "AGI Surtax", "Total"]

# Alt max policy parameters (hardcoded from YAML, matching R script)
ALT_MAX_RATE    = 0.255
ALT_MAX_QUALIFY = 1.75

# ---------------------------------------------------------------------------
# Style helpers (matching Booker template exactly)
# ---------------------------------------------------------------------------
ARIAL11       = Font(name="Arial", size=11)
ARIAL11_BOLD  = Font(name="Arial", size=11, bold=True)
ARIAL11_LINK  = Font(name="Arial", size=11, color="FF0563C1")

ALIGN_CENTER  = Alignment(horizontal="center", vertical="center")
ALIGN_H_CTR   = Alignment(horizontal="center")
ALIGN_LEFT    = Alignment(horizontal="left")
ALIGN_WRAP    = Alignment(wrap_text=True, horizontal="center", vertical="center")

FMT_COMMA = "#,##0"
FMT_PCT   = "0.00%"
FMT_1DP   = "0.0"
FMT_INT   = "0"


def std_cell(ws, row, col, value, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = ARIAL11_BOLD if bold else ARIAL11
    return c


def header_cell(ws, row, col, value, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = ARIAL11_BOLD
    c.alignment = ALIGN_WRAP if wrap else ALIGN_CENTER
    return c


def data_cell(ws, row, col, value, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = ARIAL11
    c.alignment = ALIGN_H_CTR
    if number_format:
        c.number_format = number_format
    return c


def label_cell(ws, row, col, value, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = ARIAL11_BOLD if bold else ARIAL11
    return c


def merge_and_header(ws, row, col_start, col_end, value):
    start = f"{get_column_letter(col_start)}{row}"
    end   = f"{get_column_letter(col_end)}{row}"
    ws.merge_cells(f"{start}:{end}")
    c = ws[start]
    c.value = value
    c.font  = ARIAL11_BOLD
    c.alignment = ALIGN_CENTER
    return c


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------
def read_csv_file(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def to_float(v):
    if v is None or v in ("", "NA", "Inf", "-Inf"):
        return 0.0
    return float(v)


# ---------------------------------------------------------------------------
# Footnotes
# ---------------------------------------------------------------------------
DIST_NOTE = (
    "Note: Estimate universe is nondependent tax units, including nonfilers. "
    "'Income' is measured as AGI plus: above-the-line deductions, nontaxable interest, "
    "nontaxable pension income (including OASI benefits), nondeductible capital losses, "
    "employer-side payroll taxes, and inheritances."
)

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
default_ws = wb.active
wb.remove(default_ws)

SHEETS = [
    ("Data TOC", None),
    ("T1",  "Table 1. Estimated Conventional Budgetary Effects, FY2026-2055"),
    ("F1",  f"Figure 1. Alternative Maximum Tax: Illustrative Tax Liability ({YEAR_SHOW})"),
    ("F2",  f"Figure 2. Contribution to Change in After-Tax Income by Income Group ({YEAR_SHOW})"),
    ("F3",  f"Figure 3. Contribution to Change in After-Tax Income by Age Group ({YEAR_SHOW})"),
    ("F4",  f"Figure 4. Within-Group IQR of Effective Tax Rate ({YEAR_SHOW})"),
    ("F5",  f"Figure 5. Average Tax Filing Time Burden by Income Group ({YEAR_SHOW})"),
    ("F6",  f"Figure 6. Average Effective Marginal Tax Rate on Wages by AGI Percentile ({YEAR_SHOW})"),
    ("F7",  f"Figure 7. Contribution to Change in Average Effective Marginal Tax Rate on Wages by AGI Percentile ({YEAR_SHOW})"),
]

for sname, _ in SHEETS:
    wb.create_sheet(title=sname)

# ===========================================================================
# Data TOC
# ===========================================================================
ws = wb["Data TOC"]
std_cell(ws, 1, 1, "The Working Americans' Tax Cut Act", bold=True)
std_cell(ws, 2, 1, "March 2026")
std_cell(ws, 3, 1, "The Budget Lab at Yale")
std_cell(ws, 5, 1, "Tables and Figures", bold=True)

toc_entries = [(sn, t) for sn, t in SHEETS[1:]]
for i, (sheet_id, title) in enumerate(toc_entries):
    row = 6 + i
    c = ws.cell(row=row, column=1, value=title)
    c.font = ARIAL11_LINK
    c.hyperlink = Hyperlink(ref=f"A{row}", location=f"'{sheet_id}'!A1", display=title)

# ===========================================================================
# T1 — Stacked revenue table
# ===========================================================================
print("Building T1...", flush=True)
ws = wb["T1"]

std_cell(ws, 1, 1, "Table 1. Estimated Conventional Budgetary Effects, FY2026-2055", bold=True)
std_cell(ws, 2, 1, "Subtitle: Billions of dollars", bold=True)
std_cell(ws, 3, 1, "Source: The Budget Lab calculations.", bold=True)

# Use precomputed stacked_revenue_table.csv
rev_rows = read_csv_file(os.path.join(ROOT, "stacked_revenue_table.csv"))

years = list(range(2026, 2036))

# Row 5: year headers
header_cell(ws, 5, 1, "Provision")
ws["A5"].alignment = ALIGN_LEFT
for j, yr in enumerate(years):
    c = header_cell(ws, 5, 2 + j, yr)
    c.number_format = FMT_INT

# Merged window headers in row 5
merge_and_header(ws, 5, 12, 13, "Budget Window")
merge_and_header(ws, 5, 14, 15, "Second Decade")
merge_and_header(ws, 5, 16, 17, "Third Decade")

# Row 6: sub-headers for windows
for col, val in [(12, "Dollars"), (13, "% of GDP"),
                 (14, "Dollars"), (15, "% of GDP"),
                 (16, "Dollars"), (17, "% of GDP")]:
    header_cell(ws, 6, col, val)

# Data rows (7+)
for i, r in enumerate(rev_rows):
    data_row = 7 + i
    label_cell(ws, data_row, 1, r["Provision"])
    for j, yr in enumerate(years):
        data_cell(ws, data_row, 2 + j, round(float(r[str(yr)]), 1), FMT_COMMA)
    # Budget Window
    data_cell(ws, data_row, 12, round(float(r["Budget Window"]), 1), FMT_COMMA)
    data_cell(ws, data_row, 13, float(r["Budget Window (% GDP)"]) / 100, FMT_PCT)
    # Second Decade
    data_cell(ws, data_row, 14, round(float(r["Second Decade"]), 1), FMT_COMMA)
    data_cell(ws, data_row, 15, float(r["Second Decade (% GDP)"]) / 100, FMT_PCT)
    # Third Decade
    data_cell(ws, data_row, 16, round(float(r["Third Decade"]), 1), FMT_COMMA)
    data_cell(ws, data_row, 17, float(r["Third Decade (% GDP)"]) / 100, FMT_PCT)

ws.column_dimensions["A"].width = 32
ws.column_dimensions["L"].width = 12

# ===========================================================================
# F1 — Policy illustration: Illustrative Tax Liability
# ===========================================================================
print("Building F1...", flush=True)
ws = wb["F1"]

std_cell(ws, 1, 1,
    f"Figure 1. Alternative Maximum Tax: Illustrative Tax Liability ({YEAR_SHOW})",
    bold=True)
std_cell(ws, 2, 1,
    "All income from earnings; standard deduction; CTC included for married couple",
    bold=True)
std_cell(ws, 3, 1, "Source: The Budget Lab calculations.", bold=True)

# Read tax law params for illustration
tax_law_all = read_csv_file(os.path.join(ROOT, "baseline/static/supplemental/tax_law.csv"))

def get_tl(fs):
    for r in tax_law_all:
        if int(to_float(r["year"])) == YEAR_SHOW and int(to_float(r["filing_status"])) == fs:
            return r
    return None

tl_single  = get_tl(1)
tl_married = get_tl(2)

def get_brackets_rates(tl):
    brackets, rates = [], []
    for i in range(1, 8):
        bk, rk = f"ord.brackets{i}", f"ord.rates{i}"
        if bk in tl and rk in tl:
            brackets.append(to_float(tl[bk]))
            rates.append(to_float(tl[rk]))
    return brackets, rates, to_float(tl["std.value"])

sched_s = get_brackets_rates(tl_single)
sched_m = get_brackets_rates(tl_married)

def calc_bracket_tax(agi, std_ded, brackets, rates):
    txbl = max(0, agi - std_ded)
    tax = 0.0
    for i, rate in enumerate(rates):
        top = brackets[i + 1] if i + 1 < len(brackets) else float("inf")
        tax += rate * max(0, min(txbl, top) - brackets[i])
    return tax

def get_eitc_params(tl, n_kids):
    s = str(n_kids)
    return {
        "pi_rate":   to_float(tl[f"eitc.pi_rate_{s}"]),
        "pi_end":    to_float(tl[f"eitc.pi_end_{s}"]),
        "po_thresh": to_float(tl[f"eitc.po_thresh_{s}"]),
        "po_rate":   to_float(tl[f"eitc.po_rate_{s}"]),
    }

eitc_s0 = get_eitc_params(tl_single, 0)
eitc_m2 = get_eitc_params(tl_married, 2)

ctc_per_kid   = to_float(tl_married["ctc.value_old1"])
ctc_max_ref   = to_float(tl_married["ctc.max_refund_old"])
ctc_pi_thresh = to_float(tl_married["ctc.pi_thresh"])
ctc_pi_rate   = to_float(tl_married["ctc.pi_rate"])

def calc_eitc(earnings, agi, pars):
    credit = min(pars["pi_rate"] * earnings, pars["pi_rate"] * pars["pi_end"])
    phaseout = max(0, pars["po_rate"] * (agi - pars["po_thresh"]))
    return max(0, credit - phaseout)

def calc_ctc(earnings, sec1_tax, n_kids):
    ctc_total = n_kids * ctc_per_kid
    ctc_nonref = min(ctc_total, sec1_tax)
    remaining = ctc_total - ctc_nonref
    actc_earned = ctc_pi_rate * max(0, earnings - ctc_pi_thresh)
    actc_cap = n_kids * ctc_max_ref
    ctc_ref = min(remaining, actc_earned, actc_cap)
    return ctc_nonref, ctc_ref

def build_filer_schedule(agi_seq, brackets, rates, std_ded, exempt, n_kids, eitc_pars):
    threshold = ALT_MAX_QUALIFY * exempt
    results = []
    for agi in agi_seq:
        earnings = agi
        bracket_tax = calc_bracket_tax(agi, std_ded, brackets, rates)
        eitc = calc_eitc(earnings, agi, eitc_pars)

        ctc_nr_cl, ctc_r_cl = calc_ctc(earnings, bracket_tax, n_kids)
        current_law = bracket_tax - ctc_nr_cl - ctc_r_cl - eitc

        alt_tax = ALT_MAX_RATE * max(0, agi - exempt)
        qualifies = agi < threshold
        sec1_policy = alt_tax if (qualifies and alt_tax < bracket_tax) else bracket_tax
        ctc_nr_pr, ctc_r_pr = calc_ctc(earnings, sec1_policy, n_kids)
        proposal = sec1_policy - ctc_nr_pr - ctc_r_pr - eitc

        results.append((agi, current_law, proposal))
    return results

agi_seq = list(range(0, 200001, 500))

data_single  = build_filer_schedule(agi_seq, *sched_s, exempt=46000, n_kids=0, eitc_pars=eitc_s0)
data_married = build_filer_schedule(agi_seq, *sched_m, exempt=92000, n_kids=2, eitc_pars=eitc_m2)

# Write two panels side by side (A-D single, F-I married)
# Row 4 panel labels
merge_and_header(ws, 4, 1, 4, "Single Filer, No Children")
merge_and_header(ws, 4, 6, 9, "Married Couple, Two Children")

# Row 5 headers
for col, val in [(1, "AGI"), (2, "Current Law"), (3, "Proposal"), (4, "Change")]:
    header_cell(ws, 5, col, val)
ws["A5"].alignment = ALIGN_LEFT

for col, val in [(6, "AGI"), (7, "Current Law"), (8, "Proposal"), (9, "Change")]:
    header_cell(ws, 5, col, val)
ws["F5"].alignment = ALIGN_LEFT

for idx, (agi, cl, pr) in enumerate(data_single):
    r = 6 + idx
    data_cell(ws, r, 1, agi, FMT_COMMA)
    data_cell(ws, r, 2, round(cl, 2), FMT_COMMA)
    data_cell(ws, r, 3, round(pr, 2), FMT_COMMA)
    data_cell(ws, r, 4, round(pr - cl, 2), FMT_COMMA)

for idx, (agi, cl, pr) in enumerate(data_married):
    r = 6 + idx
    data_cell(ws, r, 6, agi, FMT_COMMA)
    data_cell(ws, r, 7, round(cl, 2), FMT_COMMA)
    data_cell(ws, r, 8, round(pr, 2), FMT_COMMA)
    data_cell(ws, r, 9, round(pr - cl, 2), FMT_COMMA)

ws.column_dimensions["A"].width = 12

# ===========================================================================
# F2 — Distribution by income group
# ===========================================================================
print("Building F2...", flush=True)
ws = wb["F2"]

std_cell(ws, 1, 1,
    f"Figure 2. Contribution to Change in After-Tax Income by Income Group, {YEAR_SHOW}",
    bold=True)
std_cell(ws, 2, 1, DIST_NOTE, bold=True)
std_cell(ws, 3, 1, "Source: The Budget Lab calculations.", bold=True)

# Read distribution for both cumulative scenarios
dist_data = {}
for scn in SCENARIOS:
    rows = read_csv_file(os.path.join(ROOT, scn, "static/supplemental/distribution.csv"))
    for r in rows:
        if (int(to_float(r["year"])) == YEAR_SHOW and
            r["taxes_included"] == "iit_pr"):
            dist_data[(scn, r["group_dimension"], r["group"])] = to_float(r["pct_chg_ati"])

INCOME_MAIN   = ["Quintile 1", "Quintile 2", "Quintile 3", "Quintile 4", "Quintile 5"]
INCOME_TOPTEN = ["Top 10%", "Top 5%", "Top 1%", "Top 0.1%"]

header_cell(ws, 5, 1, "Income Group")
ws["A5"].alignment = ALIGN_LEFT
for col, val in [(2, "Alternative Maximum Tax"), (3, "AGI Surtax"), (4, "Total")]:
    header_cell(ws, 5, col, val, wrap=True)
ws.row_dimensions[5].height = 28

cur_row = 6
for g in INCOME_MAIN:
    am = dist_data.get(("alt_max", "Income", g), 0)
    sx = dist_data.get(("surtax", "Income", g), 0)
    label_cell(ws, cur_row, 1, g)
    data_cell(ws, cur_row, 2, am, FMT_PCT)
    data_cell(ws, cur_row, 3, sx - am, FMT_PCT)
    data_cell(ws, cur_row, 4, sx, FMT_PCT)
    cur_row += 1

# Separator
label_cell(ws, cur_row, 1, "Top Decile Breakout", bold=True)
cur_row += 1

for g in INCOME_TOPTEN:
    am = dist_data.get(("alt_max", "Income", g), 0)
    sx = dist_data.get(("surtax", "Income", g), 0)
    label_cell(ws, cur_row, 1, g)
    data_cell(ws, cur_row, 2, am, FMT_PCT)
    data_cell(ws, cur_row, 3, sx - am, FMT_PCT)
    data_cell(ws, cur_row, 4, sx, FMT_PCT)
    cur_row += 1

ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 25

# ===========================================================================
# F3 — Distribution by age group
# ===========================================================================
print("Building F3...", flush=True)
ws = wb["F3"]

std_cell(ws, 1, 1,
    f"Figure 3. Contribution to Change in After-Tax Income by Age Group, {YEAR_SHOW}",
    bold=True)
std_cell(ws, 2, 1, DIST_NOTE, bold=True)
std_cell(ws, 3, 1, "Source: The Budget Lab calculations.", bold=True)

AGE_GROUPS = ["29 and under", "30 - 39", "40 - 49", "50 - 64", "65+"]

header_cell(ws, 5, 1, "Age Group")
ws["A5"].alignment = ALIGN_LEFT
for col, val in [(2, "Alternative Maximum Tax"), (3, "AGI Surtax"), (4, "Total")]:
    header_cell(ws, 5, col, val, wrap=True)
ws.row_dimensions[5].height = 28

for i, g in enumerate(AGE_GROUPS):
    r = 6 + i
    am = dist_data.get(("alt_max", "Age", g), 0)
    sx = dist_data.get(("surtax", "Age", g), 0)
    label_cell(ws, r, 1, g)
    data_cell(ws, r, 2, am, FMT_PCT)
    data_cell(ws, r, 3, sx - am, FMT_PCT)
    data_cell(ws, r, 4, sx, FMT_PCT)

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 25

# ===========================================================================
# F4 — Horizontal equity (IQR of ETR)
# ===========================================================================
print("Building F4...", flush=True)
ws = wb["F4"]

std_cell(ws, 1, 1,
    f"Figure 4. Within-Group IQR of Effective Tax Rate, {YEAR_SHOW}",
    bold=True)
std_cell(ws, 2, 1, DIST_NOTE, bold=True)
std_cell(ws, 3, 1, "Source: The Budget Lab calculations.", bold=True)

horiz_rows = read_csv_file(os.path.join(ROOT, "surtax/static/supplemental/horizontal.csv"))
horiz_bl = {}
horiz_ref = {}
for r in horiz_rows:
    if int(to_float(r["year"])) == YEAR_SHOW and r["group_dimension"] in ("Income", "Overall"):
        if r["scenario"] == "baseline":
            horiz_bl[r["group"]] = to_float(r["avg_within_group_iqr"])
        else:
            horiz_ref[r["group"]] = to_float(r["avg_within_group_iqr"])

HE_GROUPS = INCOME_MAIN + ["Overall"]

header_cell(ws, 5, 1, "Income Group")
ws["A5"].alignment = ALIGN_LEFT
header_cell(ws, 5, 2, "Current Law")
header_cell(ws, 5, 3, "Proposal")

for i, g in enumerate(HE_GROUPS):
    r = 6 + i
    label_cell(ws, r, 1, g)
    data_cell(ws, r, 2, horiz_bl.get(g), FMT_PCT)
    data_cell(ws, r, 3, horiz_ref.get(g), FMT_PCT)

ws.column_dimensions["A"].width = 15

# ===========================================================================
# F5 — Time burden
# ===========================================================================
print("Building F5...", flush=True)
ws = wb["F5"]

std_cell(ws, 1, 1,
    f"Figure 5. Average Tax Filing Time Burden by Income Group, {YEAR_SHOW}",
    bold=True)
std_cell(ws, 2, 1, "Notes: Mean estimated filing time burden in hours.", bold=True)
std_cell(ws, 3, 1, "Source: The Budget Lab calculations.", bold=True)

tb_rows = read_csv_file(os.path.join(ROOT, "surtax/static/supplemental/time_burden.csv"))
tb_map = {}
for r in tb_rows:
    if int(to_float(r["year"])) == YEAR_SHOW and r["metric"] == "mean_burden":
        tb_map[r["group"]] = r

TB_GROUPS = INCOME_MAIN + ["Overall"]

header_cell(ws, 5, 1, "Income Group")
ws["A5"].alignment = ALIGN_LEFT
header_cell(ws, 5, 2, "Current Law")
header_cell(ws, 5, 3, "Proposal")

for i, g in enumerate(TB_GROUPS):
    r = 6 + i
    label_cell(ws, r, 1, g)
    if g in tb_map:
        data_cell(ws, r, 2, to_float(tb_map[g]["baseline"]), FMT_1DP)
        data_cell(ws, r, 3, to_float(tb_map[g]["reform"]),   FMT_1DP)

ws.column_dimensions["A"].width = 17
ws.column_dimensions["B"].width = 14

# ===========================================================================
# F6 & F7 — MTR by AGI percentile (requires detail file processing)
# ===========================================================================
DETAIL_COLS = ["weight", "agi", "mtr_wages1", "dep_age1", "dep_age2", "dep_age3"]


def read_detail_cols(path, cols):
    needed = set(cols)
    result = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            result.append({k: row[k] for k in needed if k in row})
    return result


def compute_mtr_by_pctile(detail_rows):
    """
    Compute weighted mean MTR by AGI percentile × parent status.
    Returns dict: {(pctile, is_parent): weighted_avg_mtr}
    """
    records = []
    for r in detail_rows:
        try:
            agi    = float(r["agi"])
            weight = float(r["weight"])
            mtr    = float(r["mtr_wages1"]) if r.get("mtr_wages1", "") not in ("", "NA") else 0.0
        except (ValueError, KeyError):
            continue
        if agi < 0:
            continue

        is_parent = False
        for da_col in ["dep_age1", "dep_age2", "dep_age3"]:
            val = r.get(da_col, "NA")
            if val not in ("", "NA"):
                try:
                    if float(val) < 18:
                        is_parent = True
                        break
                except ValueError:
                    pass

        records.append((agi, weight, mtr, is_parent))

    if not records:
        return {}

    records.sort(key=lambda x: x[0])
    total_weight = sum(r[1] for r in records)
    cumulative = 0.0

    pctile_bins = []
    bin_edges = [total_weight * p / 100.0 for p in range(1, 101)]
    bin_idx = 0

    for agi, wt, mtr, ip in records:
        cumulative += wt
        while bin_idx < 99 and cumulative > bin_edges[bin_idx]:
            bin_idx += 1
        pctile_bins.append((bin_idx + 1, wt, mtr, ip))

    sum_w = {}
    sum_wmtr = {}
    for pctile, wt, mtr, ip in pctile_bins:
        key = (pctile, ip)
        sum_w[key]    = sum_w.get(key, 0.0) + wt
        sum_wmtr[key] = sum_wmtr.get(key, 0.0) + wt * mtr

    result = {}
    for key in sum_w:
        result[key] = sum_wmtr[key] / sum_w[key] if sum_w[key] > 0 else 0.0
    return result


print("Reading detail files for F6/F7 (this may take a moment)...", flush=True)
detail_data = {}
for sc in ["baseline"] + SCENARIOS:
    path = os.path.join(ROOT, sc, "static/detail", f"{YEAR_SHOW}.csv")
    print(f"  Loading {sc}...", flush=True)
    detail_data[sc] = read_detail_cols(path, DETAIL_COLS)

print("Computing MTR by percentile...", flush=True)
mtr_pctile = {}
for sc in detail_data:
    mtr_pctile[sc] = compute_mtr_by_pctile(detail_data[sc])

# ===========================================================================
# F6 — Average MTR on Wages by AGI Percentile
# ===========================================================================
print("Building F6...", flush=True)
ws = wb["F6"]

std_cell(ws, 1, 1,
    f"Figure 6. Average Effective Marginal Tax Rate on Wages by AGI Percentile, {YEAR_SHOW}",
    bold=True)
std_cell(ws, 2, 1, "Source: The Budget Lab calculations.", bold=True)

# Row 4 merged panel headers
merge_and_header(ws, 4, 2, 3, "Non-parent")
merge_and_header(ws, 4, 4, 5, "Parent")

# Row 5 column headers
header_cell(ws, 5, 1, "AGI Percentile")
header_cell(ws, 5, 2, "Current Law")
header_cell(ws, 5, 3, "Proposal")
header_cell(ws, 5, 4, "Current Law")
header_cell(ws, 5, 5, "Proposal")

for pctile in range(1, 101):
    r = 5 + pctile
    data_cell(ws, r, 1, pctile)
    ws.cell(row=r, column=1).alignment = ALIGN_H_CTR
    for col, sc, ip in [
        (2, "baseline", False), (3, "surtax", False),
        (4, "baseline", True),  (5, "surtax", True),
    ]:
        val = mtr_pctile[sc].get((pctile, ip), None)
        data_cell(ws, r, col, val, FMT_PCT)

ws.column_dimensions["A"].width = 17
ws.column_dimensions["B"].width = 15

# ===========================================================================
# F7 — Stacked MTR decomposition by AGI Percentile
# ===========================================================================
print("Building F7...", flush=True)
ws = wb["F7"]

std_cell(ws, 1, 1,
    f"Figure 7. Contribution to Change in Average Effective Marginal Tax Rate on Wages by AGI Percentile, {YEAR_SHOW}",
    bold=True)
std_cell(ws, 2, 1, "Source: The Budget Lab calculations.", bold=True)

# Row 4 merged panel headers
merge_and_header(ws, 4, 2, 3, "Non-parent")
merge_and_header(ws, 4, 4, 5, "Parent")

# Row 5 column headers
header_cell(ws, 5, 1, "AGI Percentile")
for col, val in [
    (2, "Alternative Maximum Tax"), (3, "AGI Surtax"),
    (4, "Alternative Maximum Tax"), (5, "AGI Surtax"),
]:
    header_cell(ws, 5, col, val, wrap=True)
ws.row_dimensions[5].height = 42

def get_mtr(sc, pctile, is_parent):
    return mtr_pctile[sc].get((pctile, is_parent), 0.0)

for pctile in range(1, 101):
    r = 5 + pctile
    data_cell(ws, r, 1, pctile)
    ws.cell(row=r, column=1).alignment = ALIGN_H_CTR
    for ip_offset, is_parent in [(0, False), (2, True)]:
        bl = get_mtr("baseline", pctile, is_parent)
        am = get_mtr("alt_max",  pctile, is_parent)
        sx = get_mtr("surtax",   pctile, is_parent)

        d_am = am - bl
        d_sx = sx - am

        data_cell(ws, r, 2 + ip_offset, d_am, FMT_PCT)
        data_cell(ws, r, 3 + ip_offset, d_sx, FMT_PCT)

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 14

# ===========================================================================
# Save
# ===========================================================================
print(f"Saving to {OUT}...", flush=True)
wb.save(OUT)
print("Done.", flush=True)
