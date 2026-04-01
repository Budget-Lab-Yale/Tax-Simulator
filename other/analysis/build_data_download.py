"""
build_data_download.py
Generate TBL-Data-Booker-KYPA.xlsx for the Keep Your Pay Act analysis.
"""

import csv
import os
import sys

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = "/nfs/roberts/scratch/pi_nrs36/jar335/model_data/Tax-Simulator/v1/202603120745"
OUT  = os.path.join(ROOT, "TBL-Data-Booker-202603.xlsx")

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
ARIAL11       = Font(name="Arial", size=11)
ARIAL11_BOLD  = Font(name="Arial", size=11, bold=True)
ARIAL11_LINK  = Font(name="Arial", size=11, color="FF0563C1")

ALIGN_CENTER  = Alignment(horizontal="center", vertical="center")
ALIGN_H_CTR   = Alignment(horizontal="center")
ALIGN_LEFT    = Alignment(horizontal="left")
ALIGN_WRAP    = Alignment(wrap_text=True, horizontal="center", vertical="center")

FMT_COMMA    = "#,##0"
FMT_PCT      = "0.00%"
FMT_1DP      = "0.0"
FMT_INT      = "0"


def std_cell(ws, row, col, value, bold=False, h_align=None, v_align=None,
             number_format=None, wrap=False):
    """Write a cell with Arial 11pt and optional formatting."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = ARIAL11_BOLD if bold else ARIAL11
    al_h = h_align
    al_v = v_align
    al_w = wrap
    if al_h or al_v or al_w:
        c.alignment = Alignment(
            horizontal=al_h or "general",
            vertical=al_v or "bottom",
            wrap_text=al_w
        )
    if number_format:
        c.number_format = number_format
    return c


def header_cell(ws, row, col, value):
    """Bold, center-aligned (h+v), Arial 11."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = ARIAL11_BOLD
    c.alignment = ALIGN_CENTER
    return c


def data_cell(ws, row, col, value, number_format=None):
    """Center-aligned data cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = ARIAL11
    c.alignment = ALIGN_H_CTR
    if number_format:
        c.number_format = number_format
    return c


def label_cell(ws, row, col, value, bold=False):
    """Left-aligned label cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = ARIAL11_BOLD if bold else ARIAL11
    return c


def merge_and_header(ws, row, col_start, col_end, value):
    """Merge cells and write a bold centered header."""
    start = f"{get_column_letter(col_start)}{row}"
    end   = f"{get_column_letter(col_end)}{row}"
    ws.merge_cells(f"{start}:{end}")
    c = ws[start]
    c.value = value
    c.font  = ARIAL11_BOLD
    c.alignment = ALIGN_CENTER
    return c


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------
def read_csv(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def filter_rows(rows, **kwargs):
    out = []
    for r in rows:
        if all(r.get(k) == v for k, v in kwargs.items()):
            out.append(r)
    return out


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
# Remove default sheet; we'll add named sheets
default_ws = wb.active
wb.remove(default_ws)

# Sheet name → title for TOC
SHEETS = [
    ("Data TOC", None),
    ("T1", "Table 1. Estimated Conventional Budgetary Effects, FY2026-2055"),
    ("F1", "Figure 1. Contribution to Change in After-Tax Income by Income Group (2026)"),
    ("F2", "Figure 2. Contribution to Change in After-Tax Income by Age Group (2026)"),
    ("F3", "Figure 3. Within-Group IQR of Effective Tax Rate (2026)"),
    ("F4", "Figure 4. Average Tax Filing Time Burden by Income Group (2026)"),
    ("F5", "Figure 5. Average Effective Marginal Tax Rate on Wages by AGI Percentile (2026)"),
    ("F6", "Figure 6. Contribution to Change in Average Effective Marginal Tax Rate on Wages by AGI Percentile (2026)"),
]

for sname, _ in SHEETS:
    wb.create_sheet(title=sname)

# ===========================================================================
# Data TOC
# ===========================================================================
ws = wb["Data TOC"]

std_cell(ws, 1, 1, "The Keep Your Pay Act", bold=True)
std_cell(ws, 2, 1, "March 2026")
std_cell(ws, 3, 1, "The Budget Lab at Yale")
std_cell(ws, 5, 1, "Tables and Figures", bold=True)

toc_entries = [
    ("T1", "Table 1. Estimated Conventional Budgetary Effects, FY2026-2055"),
    ("F1", "Figure 1. Contribution to Change in After-Tax Income by Income Group (2026)"),
    ("F2", "Figure 2. Contribution to Change in After-Tax Income by Age Group (2026)"),
    ("F3", "Figure 3. Within-Group IQR of Effective Tax Rate (2026)"),
    ("F4", "Figure 4. Average Tax Filing Time Burden by Income Group (2026)"),
    ("F5", "Figure 5. Average Effective Marginal Tax Rate on Wages by AGI Percentile (2026)"),
    ("F6", "Figure 6. Contribution to Change in Average Effective Marginal Tax Rate on Wages by AGI Percentile (2026)"),
]

for i, (sheet_id, title) in enumerate(toc_entries):
    row = 6 + i
    c = ws.cell(row=row, column=1, value=title)
    c.font = ARIAL11_LINK
    c.hyperlink = Hyperlink(ref=f"A{row}", location=f"{sheet_id}!A1", display=title)

# ===========================================================================
# T1 — Stacked revenue table
# ===========================================================================
ws = wb["T1"]

std_cell(ws, 1, 1, "Table 1. Estimated Conventional Budgetary Effects, FY2026-2055", bold=True)
std_cell(ws, 2, 1, "Billions of dollars", bold=True)

# Row 4 headers
years = list(range(2026, 2036))
header_cell(ws, 4, 1, "Provision")
for j, yr in enumerate(years):
    c = header_cell(ws, 4, 2 + j, yr)
    c.number_format = FMT_INT

# Merged window headers
merge_and_header(ws, 4, 12, 13, "Budget Window")
merge_and_header(ws, 4, 14, 15, "Second Decade")
merge_and_header(ws, 4, 16, 17, "Third Decade")

# Row 5 sub-headers
for col, val in [(12, "Dollars"), (13, "% of GDP"),
                 (14, "Dollars"), (15, "% of GDP"),
                 (16, "Dollars"), (17, "% of GDP")]:
    header_cell(ws, 5, col, val)

# Data
rev_rows = read_csv(os.path.join(ROOT, "stacked_revenue_table.csv"))

for i, r in enumerate(rev_rows):
    data_row = 6 + i
    label_cell(ws, data_row, 1, r["Provision"])
    for j, yr in enumerate(years):
        data_cell(ws, data_row, 2 + j, round(float(r[str(yr)]), 1), FMT_COMMA)
    # Budget Window
    data_cell(ws, data_row, 12, round(float(r["Budget Window"]), 1), FMT_COMMA)
    data_cell(ws, data_row, 13, float(r["Budget Window (% GDP)"]) / 100, FMT_PCT)
    # Second Decade
    data_cell(ws, data_row, 14, round(float(r["Second Decade"]), 1), FMT_COMMA)
    data_cell(ws, data_row, 15, float(r["Second Decade (% GDP)"]) / 100, FMT_PCT)
    # Third Decade
    data_cell(ws, data_row, 16, round(float(r["Third Decade"]), 1), FMT_COMMA)
    data_cell(ws, data_row, 17, float(r["Third Decade (% GDP)"]) / 100, FMT_PCT)

# ===========================================================================
# Helper: build stacked ATI contribution table (F1/F2)
# ===========================================================================
SCENARIOS_STACKED = ["std", "std_eitc", "std_eitc_ctc", "std_eitc_ctc_ord"]

def read_dist_pct(scenario, year="2026", taxes="iit_pr", dim="Income"):
    path = os.path.join(ROOT, scenario, "static", "supplemental", "distribution.csv")
    rows = read_csv(path)
    filtered = filter_rows(rows, year=year, taxes_included=taxes, group_dimension=dim)
    return {r["group"]: float(r["pct_chg_ati"]) for r in filtered}


def read_dist_winners_losers(scenario, year="2026", taxes="iit_pr", dim="Income"):
    """Read share_cut.100 and share_raise.100 from the full cumulative scenario."""
    path = os.path.join(ROOT, scenario, "static", "supplemental", "distribution.csv")
    rows = read_csv(path)
    filtered = filter_rows(rows, year=year, taxes_included=taxes, group_dimension=dim)
    result = {}
    for r in filtered:
        result[r["group"]] = {
            "win":  float(r.get("share_cut.100", 0)),
            "lose": float(r.get("share_raise.100", 0)),
        }
    return result


def build_stacked_contributions(dim, groups):
    """
    Returns list of dicts: {group, std, eitc, ctc, ord, total}
    where contributions are marginal (stacked).
    """
    pct = {}
    for sc in SCENARIOS_STACKED:
        pct[sc] = read_dist_pct(sc, dim=dim)

    result = []
    for g in groups:
        v_std  = pct["std"].get(g, 0.0)
        v_eitc = pct["std_eitc"].get(g, 0.0) - v_std
        v_ctc  = pct["std_eitc_ctc"].get(g, 0.0) - pct["std_eitc"].get(g, 0.0)
        v_ord  = pct["std_eitc_ctc_ord"].get(g, 0.0) - pct["std_eitc_ctc"].get(g, 0.0)
        v_tot  = pct["std_eitc_ctc_ord"].get(g, 0.0)
        result.append({
            "group": g,
            "std":   v_std,
            "eitc":  v_eitc,
            "ctc":   v_ctc,
            "ord":   v_ord,
            "total": v_tot,
        })
    return result


def write_stacked_figure(ws, title, note, group_col_label, groups_below_topten,
                         groups_topten, separator_label="Top Decile Breakout"):
    """
    Write a stacked ATI contribution figure sheet.
    groups_below_topten: main income groups before separator
    groups_topten: groups after separator
    """
    std_cell(ws, 1, 1, title, bold=True)
    std_cell(ws, 2, 1, note, bold=True)

    header_cell(ws, 4, 1, group_col_label)
    for col, val in [(2, "Increase in Standard Deduction"), (3, "EITC Expansion"), (4, "CTC Expansion"),
                     (5, "Increase in Top Rates"), (6, "Total")]:
        header_cell(ws, 4, col, val)

    all_groups = groups_below_topten + groups_topten
    contrib    = build_stacked_contributions(
        "Income" if group_col_label == "Income Group" else "Age",
        all_groups
    )

    # Separate the two sets
    n_main = len(groups_below_topten)

    cur_row = 5
    for i, row_data in enumerate(contrib):
        if i == n_main:
            # Insert separator
            label_cell(ws, cur_row, 1, separator_label, bold=True)
            cur_row += 1
        label_cell(ws, cur_row, 1, row_data["group"])
        for col, key in [(2, "std"), (3, "eitc"), (4, "ctc"), (5, "ord"), (6, "total")]:
            data_cell(ws, cur_row, col, row_data[key], FMT_PCT)
        cur_row += 1

    # Winners/losers rows from the full cumulative scenario
    full_scn = SCENARIOS_STACKED[-1]
    dim = "Income" if group_col_label == "Income Group" else "Age"
    wl = read_dist_winners_losers(full_scn, dim=dim)
    cur_row += 1  # blank row
    label_cell(ws, cur_row, 1, "Tax cut >$100", bold=True)
    col = 2
    for g in all_groups:
        val = wl.get(g, {}).get("win", None)
        data_cell(ws, cur_row, col, val, FMT_PCT)
        col += 1
    cur_row += 1
    label_cell(ws, cur_row, 1, "Tax hike >$100", bold=True)
    col = 2
    for g in all_groups:
        val = wl.get(g, {}).get("lose", None)
        data_cell(ws, cur_row, col, val, FMT_PCT)
        col += 1


# ===========================================================================
# F1
# ===========================================================================
ws = wb["F1"]
INCOME_MAIN    = ["Quintile 1", "Quintile 2", "Quintile 3", "Quintile 4", "Quintile 5"]
INCOME_TOPTEN  = ["Top 10%", "Top 5%", "Top 1%", "Top 0.1%"]

write_stacked_figure(
    ws,
    title="Figure 1. Contribution to Change in After-Tax Income by Income Group (2026)",
    note="Percentage point contribution to change in after-tax income. Income groups defined by expanded cash income.",
    group_col_label="Income Group",
    groups_below_topten=INCOME_MAIN,
    groups_topten=INCOME_TOPTEN,
    separator_label="Top Decile Breakout",
)

# ===========================================================================
# F2
# ===========================================================================
ws = wb["F2"]
AGE_GROUPS = ["29 and under", "30 - 39", "40 - 49", "50 - 64", "65+"]

std_cell(ws, 1, 1, "Figure 2. Contribution to Change in After-Tax Income by Age Group (2026)", bold=True)
std_cell(ws, 2, 1, "Percentage point contribution to change in after-tax income. Income groups defined by expanded cash income.", bold=True)

header_cell(ws, 4, 1, "Age Group")
for col, val in [(2, "Increase in Standard Deduction"), (3, "EITC Expansion"), (4, "CTC Expansion"),
                 (5, "Increase in Top Rates"), (6, "Total")]:
    header_cell(ws, 4, col, val)

contrib_age = build_stacked_contributions("Age", AGE_GROUPS)
for i, row_data in enumerate(contrib_age):
    data_row = 5 + i
    label_cell(ws, data_row, 1, row_data["group"])
    for col, key in [(2, "std"), (3, "eitc"), (4, "ctc"), (5, "ord"), (6, "total")]:
        data_cell(ws, data_row, col, row_data[key], FMT_PCT)

# Winners/losers for age groups
wl_age = read_dist_winners_losers(SCENARIOS_STACKED[-1], dim="Age")
wl_row = 5 + len(AGE_GROUPS) + 1  # blank row gap
label_cell(ws, wl_row, 1, "Tax cut >$100", bold=True)
for j, g in enumerate(AGE_GROUPS):
    data_cell(ws, wl_row, 2 + j, wl_age.get(g, {}).get("win", None), FMT_PCT)
wl_row += 1
label_cell(ws, wl_row, 1, "Tax hike >$100", bold=True)
for j, g in enumerate(AGE_GROUPS):
    data_cell(ws, wl_row, 2 + j, wl_age.get(g, {}).get("lose", None), FMT_PCT)

# ===========================================================================
# F3 — Horizontal equity (IQR)
# ===========================================================================
ws = wb["F3"]

std_cell(ws, 1, 1, "Figure 3. Within-Group IQR of Effective Tax Rate (2026)", bold=True)
std_cell(ws, 2, 1,
    "Notes: IQR of effective tax rate within income group. "
    "Effective tax rate is individual income tax plus payroll tax liability divided by expanded cash income.",
    bold=True)

header_cell(ws, 4, 1, "Income Group")
header_cell(ws, 4, 2, "Current Law")
header_cell(ws, 4, 3, "Proposal")

horiz_rows = read_csv(os.path.join(ROOT, "std_eitc_ctc_ord", "static", "supplemental", "horizontal.csv"))
horiz_2026 = filter_rows(horiz_rows, year="2026", group_dimension="Income")

horiz_bl   = {r["group"]: float(r["avg_within_group_iqr"]) for r in horiz_2026 if r["scenario"] == "baseline"}
horiz_ref  = {r["group"]: float(r["avg_within_group_iqr"]) for r in horiz_2026 if r["scenario"] == "std_eitc_ctc_ord"}

for i, grp in enumerate(INCOME_MAIN):
    data_row = 5 + i
    label_cell(ws, data_row, 1, grp)
    data_cell(ws, data_row, 2, horiz_bl.get(grp, None), FMT_PCT)
    data_cell(ws, data_row, 3, horiz_ref.get(grp, None), FMT_PCT)

# ===========================================================================
# F4 — Time burden
# ===========================================================================
ws = wb["F4"]

std_cell(ws, 1, 1, "Figure 4. Average Tax Filing Time Burden by Income Group (2026)", bold=True)
std_cell(ws, 2, 1, "Notes: Mean estimated filing time burden in hours.", bold=True)

header_cell(ws, 4, 1, "Income Group")
header_cell(ws, 4, 2, "Current Law")
header_cell(ws, 4, 3, "Proposal")

tb_rows  = read_csv(os.path.join(ROOT, "std_eitc_ctc_ord", "static", "supplemental", "time_burden.csv"))
tb_2026  = filter_rows(tb_rows, year="2026", metric="mean_burden")
tb_map   = {r["group"]: r for r in tb_2026}

tb_groups = INCOME_MAIN + ["Overall"]
for i, grp in enumerate(tb_groups):
    data_row = 5 + i
    label_cell(ws, data_row, 1, grp)
    if grp in tb_map:
        data_cell(ws, data_row, 2, float(tb_map[grp]["baseline"]), FMT_1DP)
        data_cell(ws, data_row, 3, float(tb_map[grp]["reform"]),   FMT_1DP)

# ===========================================================================
# F5 & F6 — MTR by AGI percentile (detail files, use csv module)
# ===========================================================================

def read_detail_cols(path, cols):
    """Read only specific columns from a detail CSV. Returns list of dicts."""
    needed = set(cols)
    result = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            result.append({k: row[k] for k in needed if k in row})
    return result


def compute_mtr_by_pctile(detail_rows):
    """
    Given detail rows with weight, agi, mtr_wages1, dep_age1-3,
    compute weighted mean MTR by AGI percentile × parent status.

    Returns dict: {(pctile, is_parent): weighted_avg_mtr}
    """
    # Parse and filter agi >= 0
    records = []
    for r in detail_rows:
        try:
            agi    = float(r["agi"])
            weight = float(r["weight"])
            mtr    = float(r["mtr_wages1"]) if r["mtr_wages1"] not in ("", "NA") else 0.0
        except (ValueError, KeyError):
            continue
        if agi < 0:
            continue

        # Parent = any dep_age < 18
        is_parent = False
        for da_col in ["dep_age1", "dep_age2", "dep_age3"]:
            val = r.get(da_col, "NA")
            if val not in ("", "NA"):
                try:
                    if float(val) < 18:
                        is_parent = True
                        break
                except ValueError:
                    pass

        records.append((agi, weight, mtr, is_parent))

    if not records:
        return {}

    # Sort by AGI to compute percentile bins
    records.sort(key=lambda x: x[0])
    total_weight = sum(r[1] for r in records)
    cumulative   = 0.0

    # Assign percentile bin (1–100) based on cumulative weight
    # We'll accumulate weight and assign bin when crossing the threshold
    pctile_bins = []
    bin_edges = [total_weight * p / 100.0 for p in range(1, 101)]
    bin_idx   = 0

    for agi, wt, mtr, ip in records:
        cumulative += wt
        # Find which bin this falls into
        while bin_idx < 99 and cumulative > bin_edges[bin_idx]:
            bin_idx += 1
        pctile_bins.append((bin_idx + 1, wt, mtr, ip))  # bin_idx+1 → 1-based pctile

    # Weighted mean MTR by (pctile, parent)
    sum_w   = {}
    sum_wmtr= {}
    for pctile, wt, mtr, ip in pctile_bins:
        key = (pctile, ip)
        sum_w[key]    = sum_w.get(key, 0.0)    + wt
        sum_wmtr[key] = sum_wmtr.get(key, 0.0) + wt * mtr

    result = {}
    for key in sum_w:
        result[key] = sum_wmtr[key] / sum_w[key] if sum_w[key] > 0 else 0.0
    return result


DETAIL_COLS = ["weight", "agi", "mtr_wages1", "dep_age1", "dep_age2", "dep_age3"]

print("Reading detail files for F5/F6 (this may take a moment)...", flush=True)

detail_data = {}
for sc in ["baseline", "std", "std_eitc", "std_eitc_ctc", "std_eitc_ctc_ord"]:
    path = os.path.join(ROOT, sc, "static", "detail", "2026.csv")
    print(f"  Loading {sc}...", flush=True)
    detail_data[sc] = read_detail_cols(path, DETAIL_COLS)

print("Computing MTR by percentile...", flush=True)
mtr_pctile = {}
for sc in detail_data:
    mtr_pctile[sc] = compute_mtr_by_pctile(detail_data[sc])

# ===========================================================================
# F5
# ===========================================================================
ws = wb["F5"]

std_cell(ws, 1, 1,
    "Figure 5. Average Effective Marginal Tax Rate on Wages by AGI Percentile (2026)",
    bold=True)
std_cell(ws, 2, 1,
    "Notes: Weighted average marginal tax rate on primary earner wages within AGI percentile bin. "
    "Parent defined as having at least one dependent under 18.",
    bold=True)

# Row 3 merged group headers
merge_and_header(ws, 3, 2, 3, "Non-parent")
merge_and_header(ws, 3, 4, 5, "Parent")

# Row 4 column headers
header_cell(ws, 4, 1, "AGI Percentile")
header_cell(ws, 4, 2, "Current Law")
header_cell(ws, 4, 3, "Proposal")
header_cell(ws, 4, 4, "Current Law")
header_cell(ws, 4, 5, "Proposal")

for pctile in range(1, 101):
    data_row = 4 + pctile
    data_cell(ws, data_row, 1, pctile)
    for col, sc, is_parent in [
        (2, "baseline",          False),
        (3, "std_eitc_ctc_ord",  False),
        (4, "baseline",          True),
        (5, "std_eitc_ctc_ord",  True),
    ]:
        val = mtr_pctile[sc].get((pctile, is_parent), None)
        data_cell(ws, data_row, col, val, FMT_PCT)

# ===========================================================================
# F6
# ===========================================================================
ws = wb["F6"]

std_cell(ws, 1, 1,
    "Figure 6. Contribution to Change in Average Effective Marginal Tax Rate on Wages by AGI Percentile (2026)",
    bold=True)
std_cell(ws, 2, 1,
    "Notes: Percentage point contribution to change in weighted average marginal tax rate. "
    "Parent defined as having at least one dependent under 18.",
    bold=True)

# Row 3 merged group headers
merge_and_header(ws, 3, 2, 5, "Non-parent")
merge_and_header(ws, 3, 6, 9, "Parent")

# Row 4 column headers
header_cell(ws, 4, 1, "AGI Percentile")
for col, val in [
    (2, "Increase in Standard Deduction"), (3, "EITC Expansion"), (4, "CTC Expansion"), (5, "Increase in Top Rates"),
    (6, "Increase in Standard Deduction"), (7, "EITC Expansion"), (8, "CTC Expansion"), (9, "Increase in Top Rates"),
]:
    header_cell(ws, 4, col, val)

def get_mtr(sc, pctile, is_parent):
    return mtr_pctile[sc].get((pctile, is_parent), 0.0)

for pctile in range(1, 101):
    data_row = 4 + pctile
    data_cell(ws, data_row, 1, pctile)
    for ip_offset, is_parent in [(0, False), (4, True)]:
        bl  = get_mtr("baseline",          pctile, is_parent)
        s1  = get_mtr("std",               pctile, is_parent)
        s2  = get_mtr("std_eitc",          pctile, is_parent)
        s3  = get_mtr("std_eitc_ctc",      pctile, is_parent)
        s4  = get_mtr("std_eitc_ctc_ord",  pctile, is_parent)

        d_std  = s1 - bl
        d_eitc = s2 - s1
        d_ctc  = s3 - s2
        d_ord  = s4 - s3

        data_cell(ws, data_row, 2 + ip_offset, d_std,  FMT_PCT)
        data_cell(ws, data_row, 3 + ip_offset, d_eitc, FMT_PCT)
        data_cell(ws, data_row, 4 + ip_offset, d_ctc,  FMT_PCT)
        data_cell(ws, data_row, 5 + ip_offset, d_ord,  FMT_PCT)

# ===========================================================================
# Save
# ===========================================================================
print(f"Saving to {OUT}...", flush=True)
wb.save(OUT)
print("Done.", flush=True)
